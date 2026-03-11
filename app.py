import io
import re
from pathlib import Path
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── openpyxl 구버전 xlsx 호환 패치 (xfid → xfId) ─────────────────────────────
try:
    from openpyxl.styles import cell_style as _cs
    _orig_cs = _cs.CellStyle.__init__
    def _cs_patched(self, *a, xfid=None, **kw):
        if xfid is not None and "xfId" not in kw:
            kw["xfId"] = xfid
        _orig_cs(self, *a, **kw)
    _cs.CellStyle.__init__ = _cs_patched
except Exception:
    pass

APP_VERSION = "v2.3"
# ─────────────────────────────────────────────────────────────────────────────
# 9개 핵심 컬럼 키워드
# ─────────────────────────────────────────────────────────────────────────────
DEPT_KEYWORDS     = ["관리부서", "부서명", "부서", "팀명", "팀", "department"]
USER_KEYWORDS     = ["소유자", "사용자명", "사용자", "카드소유자", "성명", "사원명"]
DATE_KEYWORDS     = ["승인일자", "사용일", "거래일", "결제일", "일자", "날짜", "date"]
TIME_KEYWORDS     = ["승인시간", "사용시간", "거래시간", "시간", "time"]
MEMO_KEYWORDS     = ["적요", "비고", "내용", "memo", "remark"]
ACCT_NAME_KEYWORDS= ["상대계정명", "계정명", "계정"]
MERCHANT_KEYWORDS = ["가맹점명", "가맹점", "상호명", "상호", "업체명", "merchant"]
CATEGORY_KEYWORDS = ["업종명", "업종", "가맹점업종", "업태", "category"]
AMOUNT_KEYWORDS   = ["승인금액", "사용금액", "거래금액", "결제금액", "금액", "amount"]

# 취소 필터용 보조 컬럼
APPROVAL_TYPE_KEYWORDS = ["구분"]
SUPPLY_AMT_KEYWORDS    = ["공급가액"]
VAT_KEYWORDS           = ["부가세"]

DEFAULT_SUSPICIOUS_KEYWORDS = [
    "유흥", "나이트", "클럽", "룸살롱", "단란주점", "유흥주점", "소주방",
    "노래방", "가라오케", "노래클럽",
    "골프", "골프장", "골프클럽",
    "카지노", "안마", "안마시술소", "마사지", "성인",
    "명품", "루이비통", "구찌", "에르메스", "샤넬", "프라다",
]
FLAG_LABEL = {
    "주말_공휴일": "주말/공휴일",
    "심야_새벽":   "심야/새벽",
    "유흥_사치성": "유흥·사치성 업종",
    "고액_거래":   "고액 거래",
    "분할_결제":   "분할결제",
    "월한도_초과": "월 한도 초과",
}
# ─────────────────────────────────────────────────────────────────────────────
# 컬럼 자동감지
# ─────────────────────────────────────────────────────────────────────────────
def find_col(columns: list[str], keywords: list[str]) -> str | None:
    lc = [(c, c.lower().replace(" ", "")) for c in columns]
    for kw in keywords:
        kl = kw.lower().replace(" ", "")
        for col, cl in lc:
            if kl == cl:        # 정확 일치 우선
                return col
    for kw in keywords:
        kl = kw.lower().replace(" ", "")
        for col, cl in lc:
            if kl in cl:        # 부분 일치
                return col
    return None

def auto_detect(columns):
    return {
        "dept":          find_col(columns, DEPT_KEYWORDS),
        "user":          find_col(columns, USER_KEYWORDS),
        "date":          find_col(columns, DATE_KEYWORDS),
        "time":          find_col(columns, TIME_KEYWORDS),
        "memo":          find_col(columns, MEMO_KEYWORDS),
        "acct_name":     find_col(columns, ACCT_NAME_KEYWORDS),
        "merchant":      find_col(columns, MERCHANT_KEYWORDS),
        "category":      find_col(columns, CATEGORY_KEYWORDS),
        "amount":        find_col(columns, AMOUNT_KEYWORDS),
        "approval_type": find_col(columns, APPROVAL_TYPE_KEYWORDS),
        "supply_amt":    find_col(columns, SUPPLY_AMT_KEYWORDS),
        "vat":           find_col(columns, VAT_KEYWORDS),
    }

def col_idx(opts, val):
    return opts.index(val) if val and val in opts else 0

def to_none(v):
    return v if v != "(사용 안함)" else None
# ─────────────────────────────────────────────────────────────────────────────
# 날짜/시간 파싱
# ─────────────────────────────────────────────────────────────────────────────
def parse_dt(df, date_col, time_col):
    try:
        if time_col and time_col in df.columns:
            return pd.to_datetime(
                df[date_col].astype(str) + " " + df[time_col].astype(str), errors="coerce"
            )
        return pd.to_datetime(df[date_col], errors="coerce")
    except Exception:
        return pd.Series([pd.NaT] * len(df), index=df.index)

def has_time(df, date_col, time_col):
    if time_col:
        return True
    try:
        return bool(df[date_col].astype(str).dropna().head(20)
                    .str.contains(r"\d{2}:\d{2}", regex=True).any())
    except Exception:
        return False
# ─────────────────────────────────────────────────────────────────────────────
# 이상감지 함수
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_holidays():
    try:
        import holidays
        kr = holidays.KR(years=range(2015, 2031))
        return {str(d) for d in kr.keys()}
    except ImportError:
        return set()

def detect_weekend_holiday(dts, kr_hols):
    DOW = ["월","화","수","목","금","토","일"]
    valid = dts.notna()
    dow   = dts.dt.dayofweek
    ds    = dts.dt.date.astype(str)
    is_wknd = valid & (dow >= 5)
    is_hol  = valid & ~is_wknd & ds.isin(kr_hols)
    reasons = pd.Series("", index=dts.index)
    for d in (5, 6):
        m = is_wknd & (dow == d)
        if m.any():
            reasons[m] = f"주말({DOW[d]}요일)"
    reasons[is_hol] = "공휴일"
    return (is_wknd | is_hol).tolist(), reasons.tolist()

def detect_late_night(dts, sh=22, eh=6):
    valid = dts.notna()
    h = dts.dt.hour
    late = valid & ((h >= sh) | (h < eh))
    reasons = pd.Series("", index=dts.index)
    reasons[late] = dts[late].dt.strftime("심야/새벽(%H:%M)")
    return late.tolist(), reasons.tolist()

def detect_suspicious(df, merchant_col, category_col, keywords):
    if not keywords:
        return [False]*len(df), [""]*len(df)
    pat = "|".join(re.escape(k) for k in keywords)
    flags = pd.Series(False, index=df.index)
    reasons = pd.Series("", index=df.index)
    for col, label in [(category_col, "업종"), (merchant_col, "가맹점")]:
        if not col:
            continue
        m = df[col].astype(str).str.extract(f"({pat})", expand=False)
        hit = m.notna() & ~flags
        reasons[hit] = label + "주의(" + m[hit] + ")"
        flags |= m.notna()
    return flags.tolist(), reasons.tolist()

def detect_high_amount(df, amount_col, threshold):
    amt = pd.to_numeric(df[amount_col].astype(str).str.replace(",",""), errors="coerce").fillna(0)
    flags = amt >= threshold
    reasons = pd.Series("", index=df.index)
    reasons[flags] = amt[flags].apply(lambda x: f"고액거래({x:,.0f}원)")
    return flags.tolist(), reasons.tolist()

def detect_split_payment(df, merchant_col, date_col, min_count=2):
    n = len(df)
    flags, reasons = [False]*n, [""]*n
    try:
        work = df.copy()
        work["_d_"] = pd.to_datetime(df[date_col], errors="coerce").dt.date
        work["_m_"] = df[merchant_col].astype(str).str.strip()
        work["_p_"] = range(n)
        for (_, m), g in work.groupby(["_d_", "_m_"]):
            if len(g) < min_count or m in ("nan",""):
                continue
            for idx in g.index:
                p = work.loc[idx, "_p_"]
                flags[p] = True
                reasons[p] = f"분할결제({len(g)}회/동일일)"
    except Exception:
        pass
    return flags, reasons

def detect_monthly_limit(df, amount_col, user_col, date_col, limit):
    flags = pd.Series(False, index=df.index)
    reasons = pd.Series("", index=df.index)
    try:
        amt  = pd.to_numeric(df[amount_col].astype(str).str.replace(",",""), errors="coerce").fillna(0)
        dt   = pd.to_datetime(df[date_col], errors="coerce")
        ym   = dt.dt.to_period("M").astype(str)
        user = df[user_col].astype(str).str.strip()
        w    = pd.DataFrame({"u": user, "ym": ym, "a": amt})
        ms   = w.groupby(["u","ym"])["a"].transform("sum")
        flags = (ms >= limit) & dt.notna()
        reasons[flags] = (
            user[flags] + "/" + ym[flags] + " 월합계("
            + ms[flags].apply(lambda x: f"{x:,.0f}원") + ")"
        )
    except Exception:
        pass
    return flags.tolist(), reasons.tolist()
# ─────────────────────────────────────────────────────────────────────────────
# 엑셀 그룹형 출력
# ─────────────────────────────────────────────────────────────────────────────
def write_grouped_excel(buf, export_df, cancel_df, user_col, amount_col,
                        supply_col, vat_col, export_cols):
    GREY  = PatternFill("solid", fgColor="D9D9D9")
    SUB   = PatternFill("solid", fgColor="F2F2F2")
    RED_F = PatternFill("solid", fgColor="FFD7D7")
    YEL_F = PatternFill("solid", fgColor="FFFFD7")
    HDR_F = PatternFill("solid", fgColor="4472C4")
    RED_FONT  = Font(color="FF0000")
    BOLD_FONT = Font(bold=True)
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    num_cols = {c for c in [amount_col, supply_col, vat_col] if c}

    def row_vals(r):
        return [("" if pd.isna(r.get(c,"")) else r.get(c,"")) for c in export_cols]

    def write_header(ws):
        ws.append(export_cols)
        rn = ws.max_row
        for i in range(1, len(export_cols)+1):
            c = ws.cell(rn, i)
            c.font = HDR_FONT
            c.fill = HDR_F
            c.alignment = Alignment(horizontal="center")

    def write_nm(ws, owner):
        ws.append([f"NM_OWNER: {owner}"] + [""]*max(0, len(export_cols)-1))
        rn = ws.max_row
        if len(export_cols) > 1:
            ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=len(export_cols))
        ws.cell(rn, 1).fill = GREY
        ws.cell(rn, 1).font = BOLD_FONT

    def write_row(ws, row, is_cancel=False):
        ws.append(row_vals(row))
        rn = ws.max_row
        risk = row.get("위험점수", 0)
        if is_cancel:
            for c in ws[rn]: c.font = RED_FONT
        elif risk >= 2:
            for c in ws[rn]: c.fill = RED_F
        elif risk == 1:
            for c in ws[rn]: c.fill = YEL_F

    def write_subtotal(ws, grp):
        sub = [""]*len(export_cols)
        for col in num_cols:
            if col in export_cols:
                i = export_cols.index(col)
                try:
                    sub[i] = pd.to_numeric(
                        grp[col].astype(str).str.replace(",",""), errors="coerce"
                    ).fillna(0).sum()
                except Exception:
                    pass
        ws.append(sub)
        rn = ws.max_row
        for c in ws[rn]:
            c.fill = SUB
            c.font = BOLD_FONT

    def write_sheet(ws, approved, cancelled):
        write_header(ws)
        if user_col and user_col in approved.columns and len(approved) > 0:
            for owner in approved[user_col].dropna().unique():
                grp  = approved[approved[user_col] == owner]
                cgrp = (cancelled[cancelled[user_col] == owner]
                        if len(cancelled) > 0 and user_col in cancelled.columns
                        else pd.DataFrame())
                write_nm(ws, owner)
                for _, r in grp.iterrows():
                    write_row(ws, r)
                for _, r in cgrp.iterrows():
                    write_row(ws, r, is_cancel=True)
                write_subtotal(ws, grp)
        else:
            for _, r in approved.iterrows():
                write_row(ws, r)

    wb = Workbook()

    ws_all = wb.active
    ws_all.title = "전체결과"
    write_sheet(ws_all, export_df, cancel_df)

    ws_flag = wb.create_sheet("이상의심")
    flagged = (export_df[export_df["위험점수"] > 0]
               if "위험점수" in export_df.columns else export_df)
    write_sheet(ws_flag, flagged, pd.DataFrame())

    ws_high = wb.create_sheet("고위험")
    high = (export_df[export_df["위험점수"] >= 2]
            if "위험점수" in export_df.columns else pd.DataFrame(columns=export_cols))
    write_sheet(ws_high, high, pd.DataFrame())

    ws_sum = wb.create_sheet("요약")
    ws_sum.append(["구분", "값"])
    total   = len(export_df)
    flagged_n = int((export_df["위험점수"] > 0).sum()) if "위험점수" in export_df.columns else 0
    high_n    = int((export_df["위험점수"] >= 2).sum()) if "위험점수" in export_df.columns else 0
    for row in [
        ("총 승인건수", total),
        ("취소 거래(제외)", len(cancel_df)),
        ("이상 의심 건수", flagged_n),
        ("고위험 건수", high_n),
        ("이상 비율", f"{flagged_n/total*100:.1f}%" if total else "0%"),
    ]:
        ws_sum.append(list(row))

    wb.save(buf)
# ─────────────────────────────────────────────────────────────────────────────
# 메인 앱
# ─────────────────────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="법인카드 이상징후 스크리닝",
        page_icon="🔍",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.title(f"🔍 법인카드 이상징후 스크리닝  `{APP_VERSION}`")
    st.caption("법인카드 내역을 업로드하면 이상징후를 자동 분석합니다.")

    # ── 사이드바 설정 ──────────────────────────────────────────────────────────
    with st.sidebar:
        st.header("⚙️ 탐지 설정")
        use_weekend = st.checkbox("주말/공휴일 사용 탐지", value=True)
        use_late = st.checkbox("심야/새벽 사용 탐지", value=True)
        if use_late:
            late_start = st.slider("심야 시작 시간", 18, 23, 22)
            late_end   = st.slider("새벽 종료 시간", 1, 9, 6)
        else:
            late_start, late_end = 22, 6

        use_suspicious = st.checkbox("유흥·사치성 업종 탐지", value=True)
        custom_kw = st.text_area("추가 탐지 키워드 (줄바꿈 구분)",
                                  placeholder="예:\n뷔페\n리조트", height=80)

        use_high = st.checkbox("고액 거래 탐지", value=False)
        high_thr = st.number_input("기준 금액 (원)", 0, value=300000, step=10000,
                                    format="%d") if use_high else 300000

        use_split = st.checkbox("분할결제 탐지", value=True)
        split_min = st.slider("동일일 동일가맹점 최소 횟수", 2, 5, 2) if use_split else 2

        use_limit = st.checkbox("인당 월 한도 초과 탐지", value=False)
        monthly_limit = st.number_input("월 한도 기준 (원)", 0, value=500000, step=100000,
                                         format="%d") if use_limit else 500000

        st.divider()
        st.subheader("🏦 더존 iUERP 옵션")
        exclude_cancel = st.checkbox("취소 거래 제외", value=True)

    keywords = DEFAULT_SUSPICIOUS_KEYWORDS.copy()
    if custom_kw:
        keywords.extend(k.strip() for k in custom_kw.splitlines() if k.strip())

    # ── 파일 업로드 ────────────────────────────────────────────────────────────
    st.header("1️⃣ 파일 업로드")
    uploaded = st.file_uploader("법인카드 내역 파일 (xlsx / xls / csv)",
                                 type=["xlsx","xls","csv"])
    if uploaded is None:
        st.info("👆 파일을 업로드하면 분석이 시작됩니다.")
        return

    try:
        if uploaded.name.lower().endswith(".csv"):
            for enc in ("utf-8-sig","utf-8","cp949","euc-kr"):
                try:
                    df = pd.read_csv(uploaded, encoding=enc); uploaded.seek(0); break
                except UnicodeDecodeError:
                    uploaded.seek(0)
        elif uploaded.name.lower().endswith(".xls"):
            xl = pd.ExcelFile(uploaded, engine="xlrd")
            sheet = (st.selectbox("시트 선택", xl.sheet_names)
                     if len(xl.sheet_names) > 1 else xl.sheet_names[0])
            df = pd.read_excel(uploaded, sheet_name=sheet, engine="xlrd")
        else:
            xl = pd.ExcelFile(uploaded, engine="openpyxl")
            sheet = (st.selectbox("시트 선택", xl.sheet_names)
                     if len(xl.sheet_names) > 1 else xl.sheet_names[0])
            uploaded.seek(0)
            df = pd.read_excel(uploaded, sheet_name=sheet, engine="openpyxl")
    except Exception as e:
        st.error(f"파일 읽기 오류: {e}"); return

    if df.empty:
        st.error("데이터가 없습니다."); return

    # ── 더존 iUERP 전처리: NM_OWNER 헤더 행 / 소계 행 제거 ─────────────────
    _before = len(df)
    # NM_OWNER 행: 어느 컬럼이든 "NM_OWNER:" 문자열을 포함
    _nm_mask = df.apply(lambda r: r.astype(str).str.contains("NM_OWNER:", na=False).any(), axis=1)
    # 소계 행: 모든 컬럼이 NaN 이거나 빈 문자열 (숫자 합계 행)
    _blank_mask = df.apply(
        lambda r: r.map(lambda v: pd.isna(v) or str(v).strip() in ("", "nan")).all(), axis=1
    )
    df = df[~(_nm_mask | _blank_mask)].reset_index(drop=True)
    _removed = _before - len(df)
    if _removed > 0:
        st.caption(f"ℹ️ iUERP 그룹 헤더·소계 행 {_removed}개 자동 제거")

    st.success(f"✅ 로드 완료: **{len(df):,}건** · **{len(df.columns)}개** 컬럼")
    with st.expander("📄 원본 데이터 미리보기 (상위 10행)"):
        st.dataframe(df.head(10), use_container_width=True, hide_index=True)

    # ── 컬럼 매핑 ─────────────────────────────────────────────────────────────
    st.header("2️⃣ 컬럼 매핑")
    auto = auto_detect(df.columns.tolist())
    opts = ["(사용 안함)"] + df.columns.tolist()

    # 더존 iUERP 기본 프리셋
    _PRESET = {
        "cm_dept": "관리부서", "cm_user": "소유자",
        "cm_date": "승인일자", "cm_time": "승인시간",
        "cm_memo": "적요",     "cm_acct": "계정명",
        "cm_merch": "가맹점",  "cm_cat": "업종",
        "cm_amt": "승인금액",
        "cm_aptype": "구분",   "cm_supply": "공급가액", "cm_vat": "부가세",
    }

    detected = sum(1 for v in auto.values() if v)
    bc = st.columns([3,1])
    with bc[0]:
        pct = detected / len(auto)
        if pct >= 0.7:
            st.success(f"더존 iU 형식 감지 · {detected}/{len(auto)} 컬럼 자동 인식")
        elif pct >= 0.4:
            st.warning(f"일부 컬럼 미감지 · {detected}/{len(auto)} · 아래에서 확인하세요")
        else:
            st.error(f"컬럼 자동 감지 실패 · {detected}/{len(auto)} · 수동 설정 필요")
    with bc[1]:
        if st.button("더존 iU 기본형식 적용", use_container_width=True):
            for k, v in _PRESET.items():
                if v in opts:
                    st.session_state[k] = v
            st.rerun()

    def sv(key, auto_val):
        return st.session_state.get(key, auto_val)

    with st.expander("컬럼 매핑 확인 / 수정", expanded=(pct < 0.7)):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**핵심 9개 컬럼**")
            sel_dept  = st.selectbox("A 관리부서", opts, key="cm_dept",  index=col_idx(opts, sv("cm_dept",  auto["dept"])))
            sel_user  = st.selectbox("B 소유자",   opts, key="cm_user",  index=col_idx(opts, sv("cm_user",  auto["user"])))
            sel_date  = st.selectbox("C 승인일자 *", opts, key="cm_date", index=col_idx(opts, sv("cm_date",  auto["date"])))
            sel_time  = st.selectbox("D 승인시간", opts, key="cm_time",  index=col_idx(opts, sv("cm_time",  auto["time"])))
            sel_memo  = st.selectbox("E 적요",     opts, key="cm_memo",  index=col_idx(opts, sv("cm_memo",  auto["memo"])))
            sel_acct  = st.selectbox("F 계정명",   opts, key="cm_acct",  index=col_idx(opts, sv("cm_acct",  auto["acct_name"])))
            sel_merch = st.selectbox("G 가맹점",   opts, key="cm_merch", index=col_idx(opts, sv("cm_merch", auto["merchant"])))
            sel_cat   = st.selectbox("H 업종",     opts, key="cm_cat",   index=col_idx(opts, sv("cm_cat",   auto["category"])))
            sel_amt   = st.selectbox("I 승인금액", opts, key="cm_amt",   index=col_idx(opts, sv("cm_amt",   auto["amount"])))
        with c2:
            st.markdown("**보조 컬럼 (소계·취소 필터용)**")
            sel_aptype = st.selectbox("구분 (승인/취소)", opts, key="cm_aptype",
                                       index=col_idx(opts, sv("cm_aptype", auto["approval_type"])))
            sel_supply = st.selectbox("공급가액",         opts, key="cm_supply",
                                       index=col_idx(opts, sv("cm_supply", auto["supply_amt"])))
            sel_vat    = st.selectbox("부가세",           opts, key="cm_vat",
                                       index=col_idx(opts, sv("cm_vat",    auto["vat"])))

    dept_col    = to_none(sel_dept)
    user_col    = to_none(sel_user)
    date_col    = to_none(sel_date)
    time_col    = to_none(sel_time)
    memo_col    = to_none(sel_memo)
    acct_col    = to_none(sel_acct)
    merchant_col= to_none(sel_merch)
    category_col= to_none(sel_cat)
    amount_col  = to_none(sel_amt)
    aptype_col  = to_none(sel_aptype)
    supply_col  = to_none(sel_supply)
    vat_col_    = to_none(sel_vat)

    if not date_col:
        st.warning("날짜 컬럼을 선택해야 분석을 진행할 수 있습니다.")
        return

    # ── 스크리닝 실행 ──────────────────────────────────────────────────────────
    st.header("3️⃣ 스크리닝 실행")
    run = st.button("🔍 이상징후 스크리닝 시작", type="primary", use_container_width=True)

    fkey = f"{uploaded.name}_{uploaded.size}"
    if st.session_state.get("_fkey") != fkey:
        st.session_state.pop("_cache", None)
        st.session_state["_fkey"] = fkey

    if run:
        prog = st.progress(0, "분석 준비 중...")
        work = df.copy()

        # 취소 거래 분리
        cancel_df = pd.DataFrame()
        if exclude_cancel and aptype_col and aptype_col in work.columns:
            mask = work[aptype_col].astype(str).str.strip().isin(["취소","취소(전체)","CANCEL"])
            cancel_df = work[mask].copy()
            work = work[~mask].reset_index(drop=True)
        if len(cancel_df):
            st.info(f"ℹ️ 취소 거래 **{len(cancel_df):,}건** 제외")

        dts = parse_dt(work, date_col, time_col)
        work["_dt_"] = dts
        flag_cols = []

        if use_weekend:
            prog.progress(15, "공휴일 로드 중...")
            hols = load_holidays()
            prog.progress(30, "주말/공휴일 탐지 중...")
            f, r = detect_weekend_holiday(dts, hols)
            work["주말_공휴일"] = f; work["주말_공휴일_사유"] = r
            flag_cols.append("주말_공휴일")

        if use_late:
            prog.progress(45, "심야/새벽 탐지 중...")
            if has_time(work, date_col, time_col):
                f, r = detect_late_night(dts, late_start, late_end)
                work["심야_새벽"] = f; work["심야_새벽_사유"] = r
                flag_cols.append("심야_새벽")
            else:
                st.info("시간 정보 없음 — 심야/새벽 탐지 건너뜀")

        if use_suspicious and (merchant_col or category_col):
            prog.progress(60, "유흥·사치성 탐지 중...")
            f, r = detect_suspicious(work, merchant_col, category_col, keywords)
            work["유흥_사치성"] = f; work["유흥_사치성_사유"] = r
            flag_cols.append("유흥_사치성")

        if use_high and amount_col:
            prog.progress(75, "고액 거래 탐지 중...")
            f, r = detect_high_amount(work, amount_col, int(high_thr))
            work["고액_거래"] = f; work["고액_거래_사유"] = r
            flag_cols.append("고액_거래")

        if use_split and merchant_col:
            prog.progress(85, "분할결제 탐지 중...")
            f, r = detect_split_payment(work, merchant_col, date_col, split_min)
            work["분할_결제"] = f; work["분할_결제_사유"] = r
            flag_cols.append("분할_결제")

        if use_limit and amount_col and user_col:
            prog.progress(93, "월 한도 초과 탐지 중...")
            f, r = detect_monthly_limit(work, amount_col, user_col, date_col, int(monthly_limit))
            work["월한도_초과"] = f; work["월한도_초과_사유"] = r
            flag_cols.append("월한도_초과")

        prog.progress(98, "결과 집계 중...")
        work["위험점수"] = work[flag_cols].sum(axis=1).astype(int) if flag_cols else 0
        work["위험등급"] = work["위험점수"].map(
            lambda s: "🔴 위험" if s >= 2 else ("🟡 주의" if s == 1 else "🟢 정상")
        )
        rcols = [c for c in work.columns if c.endswith("_사유")]
        work["이상사유"] = work[rcols].apply(
            lambda row: " | ".join(v for v in row if v and str(v) not in ("","nan")), axis=1
        )
        prog.progress(100, "완료!")
        prog.empty()

        st.session_state["_cache"] = {
            "work": work, "cancel": cancel_df, "flag_cols": flag_cols,
            "cols": {
                "dept": dept_col, "user": user_col, "date": date_col,
                "time": time_col, "memo": memo_col, "acct": acct_col,
                "merchant": merchant_col, "category": category_col,
                "amount": amount_col, "aptype": aptype_col,
                "supply": supply_col, "vat": vat_col_,
            },
        }

    if "_cache" not in st.session_state:
        return

    # ── 결과 복원 ──────────────────────────────────────────────────────────────
    cache       = st.session_state["_cache"]
    result      = cache["work"]
    cancel_df   = cache.get("cancel", pd.DataFrame())
    flag_cols   = cache["flag_cols"]
    cols        = cache["cols"]
    dept_col    = cols["dept"];    user_col = cols["user"]
    date_col    = cols["date"];    time_col = cols["time"]
    memo_col    = cols["memo"];    acct_col = cols["acct"]
    merchant_col= cols["merchant"]; category_col = cols["category"]
    amount_col  = cols["amount"];  supply_col = cols["supply"]
    vat_col_    = cols["vat"]
    dts         = result["_dt_"]

    # ── 엑셀 다운로드 ─────────────────────────────────────────────────────────
    st.header("4️⃣ 분석 결과")

    drop = (list(flag_cols)
            + [c for c in result.columns if c.endswith("_사유")]
            + ["_dt_"])
    export = result.drop(columns=drop, errors="ignore")

    if len(cancel_df) > 0:
        cexp = cancel_df.drop(columns=[c for c in drop if c in cancel_df.columns], errors="ignore")
        for col in ["위험등급","이상사유","위험점수"]:
            if col not in cexp.columns:
                cexp[col] = "취소"
    else:
        cexp = pd.DataFrame()

    desired = [
        dept_col, user_col, date_col, time_col,
        memo_col, acct_col, merchant_col, category_col, amount_col,
        supply_col, vat_col_,
        "이상사유", "위험등급",
    ]
    seen: set = set()
    ecols: list = []
    for c in desired:
        if c and c in export.columns and c not in seen:
            ecols.append(c); seen.add(c)
    skip = set(drop) | {"위험점수"}
    for c in export.columns:
        if c not in seen and c not in skip:
            ecols.append(c); seen.add(c)

    buf = io.BytesIO()
    write_grouped_excel(buf, export, cexp, user_col, amount_col,
                        supply_col, vat_col_, ecols)
    buf.seek(0)
    st.download_button(
        label="📥 분석 결과 엑셀 다운로드",
        data=buf.getvalue(),
        file_name=f"{Path(uploaded.name).stem}_스크리닝.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # ── 상세 결과 테이블 ───────────────────────────────────────────────────────
    st.subheader("📋 상세 결과")
    mn = dts.dropna().dt.date.min() if dts.notna().any() else None
    mx = dts.dropna().dt.date.max() if dts.notna().any() else None
    date_range = None
    if mn and mx and mn != mx:
        date_range = st.date_input("📅 기간 필터", value=(mn,mx), min_value=mn, max_value=mx)

    fa, fb = st.columns([1,2])
    with fa:
        show_f = st.selectbox("표시 범위",
                               ["전체","이상 의심만 (주의+위험)","고위험만 (🔴 위험)"])
    with fb:
        type_opts = [FLAG_LABEL.get(c,c) for c in flag_cols]
        type_f = st.multiselect("이상징후 유형 필터", options=type_opts)

    disp = result.copy()
    if date_range and len(date_range) == 2:
        disp = disp[(disp["_dt_"].dt.date >= date_range[0]) &
                    (disp["_dt_"].dt.date <= date_range[1])]
    if show_f == "이상 의심만 (주의+위험)":
        disp = disp[disp["위험점수"] > 0]
    elif show_f == "고위험만 (🔴 위험)":
        disp = disp[disp["위험점수"] >= 2]
    if type_f:
        rev = {v:k for k,v in FLAG_LABEL.items()}
        tf  = [rev.get(t,t) for t in type_f if rev.get(t,t) in disp.columns]
        if tf:
            disp = disp[disp[tf].any(axis=1)]

    # 표시 컬럼: 위험등급·이상사유 + 9개 핵심 컬럼
    show_cols = ["위험등급","이상사유"]
    for c in [dept_col, user_col, date_col, time_col,
              memo_col, acct_col, merchant_col, category_col, amount_col]:
        if c and c not in show_cols:
            show_cols.append(c)
    show_cols = [c for c in show_cols if c in disp.columns]

    MAX_ROWS = 500
    total_disp = len(disp)
    if total_disp > MAX_ROWS:
        st.warning(f"⚡ 상위 {MAX_ROWS:,}건만 표시 (전체 {total_disp:,}건은 엑셀 다운로드)")
        disp = disp.head(MAX_ROWS)

    def _grade_bg(s):
        return ["background-color:#fde8e8" if "위험" in str(v)
                else "background-color:#fef9e7" if "주의" in str(v)
                else "" for v in s]

    styled = disp[show_cols].style
    if "위험등급" in show_cols:
        styled = styled.apply(_grade_bg, subset=["위험등급"])
    if amount_col and amount_col in show_cols:
        styled = styled.format(
            {amount_col: lambda x: f"{float(str(x).replace(',','')):,.0f}"
             if str(x) not in ("nan","","0") else "-"},
            na_rep="-"
        )
    st.caption(f"표시 {min(total_disp,MAX_ROWS):,}건 / 전체 {total_disp:,}건")
    st.dataframe(styled, use_container_width=True, height=420, hide_index=True)

if __name__ == "__main__":
    main()
